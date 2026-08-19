from pathlib import Path

from openpyxl import Workbook, load_workbook

from ehrm.core.error_catalog import configure_error_messages
from ehrm.modules.rights_statement.excel_models import ItemResult
from ehrm.modules.rights_statement.result_workbook import ResultWorkbookWriter


def test_appends_row_level_failure_reason_without_changing_source(
    tmp_path: Path,
) -> None:
    configure_error_messages(Path("config/error_messages.toml"))
    source = tmp_path / "人员.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["单位", "姓名"])
    sheet.append(["测试单位", "成功人员"])
    sheet.append(["测试单位", "失败人员"])
    workbook.save(source)
    workbook.close()

    result = ResultWorkbookWriter().write(
        source,
        tmp_path / "output",
        [
            ItemResult(2, True, "SUCCESS", "下载成功"),
            ItemResult(3, False, "EMPLOYEE_NOT_FOUND", "没有查询到人员"),
        ],
    )

    original = load_workbook(source, read_only=True)
    exported = load_workbook(result, read_only=True)
    try:
        assert original.active.max_column == 2
        assert exported.active.cell(1, 3).value == "失败原因"
        assert exported.active.cell(2, 3).value in (None, "")
        assert exported.active.cell(3, 3).value == "未查询到符合条件的人员"
        assert exported.active.cell(1, 4).value == "ERP上传结果"
        assert exported.active.cell(2, 4).value == "未执行"
        assert exported.active.cell(1, 5).value == "ERP失败原因"
    finally:
        original.close()
        exported.close()
