from pathlib import Path

from openpyxl import load_workbook

from ehrm.gui.template_service import RightsStatementTemplateService
from ehrm.modules.rights_statement.excel_models import EmployeeRecord


def test_template_contains_required_columns_and_insurance_dropdown(
    tmp_path: Path,
) -> None:
    destination = RightsStatementTemplateService().write(tmp_path / "模板.xlsx")

    workbook = load_workbook(destination)
    try:
        sheet = workbook["人员导入"]
        assert tuple(cell.value for cell in sheet[1]) == (
            "任务编号",
            "单位",
            "部门",
            "姓名",
            "身份证",
            "险种",
            "开始时间",
            "结束时间",
        )
        assert sheet["A2"].number_format == "@"
        assert sheet["E2"].number_format == "@"
        validations = list(sheet.data_validations.dataValidation)
        assert len(validations) == 1
        assert validations[0].formula1 == '"养老,工伤,失业"'
        assert "F2:F1001" in str(validations[0].sqref)
        assert "填写说明" in workbook.sheetnames
    finally:
        workbook.close()


def test_erp_records_can_be_written_as_execution_source(tmp_path: Path) -> None:
    destination = RightsStatementTemplateService().write_records(
        tmp_path / "ERP申请解析数据.xlsx",
        [
            EmployeeRecord(
                row_number=2,
                unit="南京南化建设有限公司",
                department="技术中心",
                name="张三",
                identity_number="320101199001011234",
                insurance_type="养老",
                start_month="2025-07",
                end_month="2026-07",
                task_number="RLSQ20260818-0004",
                print_group_id="RLSQ20260818-0004-G01",
                print_group_sequence=1,
                resolved_print_mode="combined",
            )
        ],
    )

    workbook = load_workbook(destination, data_only=True)
    try:
        sheet = workbook["ERP申请解析数据"]
        assert sheet.max_row == 2
        assert [cell.value for cell in sheet[1]] == [
            "任务编号",
            "打印组",
            "打印方式",
            "单位",
            "部门",
            "姓名",
            "身份证",
            "险种",
            "开始时间",
            "结束时间",
        ]
        assert [cell.value for cell in sheet[2]] == [
            "RLSQ20260818-0004",
            "组1",
            "合并打印",
            "南京南化建设有限公司",
            "技术中心",
            "张三",
            "320101199001011234",
            "养老",
            "2025-07",
            "2026-07",
        ]
    finally:
        workbook.close()


def test_edited_excel_records_use_the_original_eight_column_structure(
    tmp_path: Path,
) -> None:
    destination = RightsStatementTemplateService().write_records(
        tmp_path / "人员执行数据.xlsx",
        [
            EmployeeRecord(
                row_number=2,
                unit="测试单位",
                department="项目部",
                name="张三",
                identity_number="320101199001011234",
                insurance_type="养老",
                start_month="2025-01",
                end_month="2025-06",
                task_number="RLSQ-001",
            )
        ],
        include_print_groups=False,
    )

    workbook = load_workbook(destination, data_only=True)
    try:
        sheet = workbook["人员执行数据"]
        assert [cell.value for cell in sheet[1]] == list(
            RightsStatementTemplateService.HEADERS
        )
        assert sheet.max_column == 8
    finally:
        workbook.close()
