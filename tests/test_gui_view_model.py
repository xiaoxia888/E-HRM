import base64
import json
import logging
import os
from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook


os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
pytest.importorskip("PySide6")

from PySide6.QtCore import QMetaObject, QObject, QPointF, Qt, QUrl
from PySide6.QtGui import QGuiApplication, QPainter, QPdfWriter
from PySide6.QtQml import QQmlApplicationEngine
from PySide6.QtQuick import QQuickItem
from PySide6.QtTest import QTest

from ehrm.core.auth_repository import AuthenticationRepository, SystemType
from ehrm.core.settings import load_settings
from ehrm.gui import view_model as view_model_module
from ehrm.gui import nocobase_connection_worker as nocobase_connection_module
from ehrm.gui.nocobase_connection_worker import NocoBaseConnectionWorker
from ehrm.gui.view_model import DesktopViewModel
from ehrm.modules.nocobase.models import (
    NocoBaseCredentials,
    NocoBaseLoginResult,
    NocoBasePageMeta,
    NocoBaseRelatedPerson,
    NocoBaseRightsApplication,
    NocoBaseRightsApplicationDetail,
    NocoBaseRightsApplicationPage,
    NocoBaseTokenClaims,
    NocoBaseUser,
)
from ehrm.modules.rights_statement.excel_models import (
    ExcelRunResult,
    ExcelTaskRequest,
    ExportMode,
    ItemResult,
)


def _view_model(tmp_path: Path) -> DesktopViewModel:
    settings = load_settings(
        Path("config/settings.toml"),
        data_root=tmp_path / "runtime",
    )
    return DesktopViewModel(
        settings,
        logging.getLogger("test.gui"),
        start_worker=False,
    )


def test_import_failure_emits_specific_row_details(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["单位", "部门", "姓名", "身份证", "险种", "开始时间", "结束时间", "任务编号"])
    sheet.append(["测试单位", "信息部", "张三", "错误证件号", "医疗", "2025-01", "2025-06", "RLSQ20260819-0001"])
    path = tmp_path / "invalid.xlsx"
    workbook.save(path)

    view_model = _view_model(tmp_path)
    failures: list[tuple[str, str]] = []
    view_model.validationFailed.connect(
        lambda summary, details: failures.append((summary, details))
    )

    view_model.importExcel(QUrl.fromLocalFile(str(path)))

    assert application is not None
    assert failures
    assert failures[0][0] == "Excel 数据校验失败"
    assert "第 2 行" in failures[0][1]
    assert "身份证格式错误" in failures[0][1]
    assert not view_model.imported


def test_successful_import_updates_preview_and_uses_safe_default(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["单位", "部门", "姓名", "身份证", "险种", "开始时间", "结束时间", "任务编号"])
    sheet.append(["测试单位", "信息部", "张三", "320101199001011234", "养老", "2025-01", "2025-06", "RLSQ20260819-0001"])
    sheet.append(["测试单位", "人事部", "李四", "320101199002021235", "养老", "2025-01", "2025-06", "RLSQ20260819-0001"])
    path = tmp_path / "valid.xlsx"
    workbook.save(path)

    view_model = _view_model(tmp_path)
    view_model.importExcel(QUrl.fromLocalFile(str(path)))

    assert application is not None
    assert view_model.exportMode == "individual"
    assert view_model.peopleCount == 2
    assert view_model.conditionCount == 1
    assert view_model.expectedPdfCount == 2
    assert view_model.batchExpectedPdfCount == 1
    assert view_model.records[0]["identity"] == "320101199001011234"
    assert view_model.records[0]["taskNumber"] == "RLSQ20260819-0001"


def test_erp_extraction_result_is_previewed_with_default_pension_insurance(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    view_model._on_erp_task_extraction_completed(
        {
            "summary": {
                "tasks_total": 1,
                "tasks_processed": 1,
                "tasks_failed": 0,
                "people_extracted": 1,
                "stopped": False,
            },
            "tasks": [
                {
                    "task_number": "RLSQ20260820-0001",
                    "department": "技术中心",
                }
            ],
            "rights_statement_requests": [
                {
                    "task_number": "RLSQ20260820-0001",
                    "name": "张三",
                    "social_security_number": None,
                    "start_month": "2025-08",
                    "end_month": "2026-07",
                }
            ],
        }
    )

    assert application is not None
    assert view_model.hasRecords
    assert not view_model.imported
    assert view_model.fileSummary == "ERP 申请解析结果 · 1 条人员记录"
    assert view_model.recordIssueCount == 1
    assert view_model.recordIssues[0]["code"] == "IDENTITY_MATCH_PENDING"
    assert view_model.records[0] == {
        "status": "通过",
        "rowNumber": 2,
        "rowStatus": "error",
        "rowStatusLabel": "错误",
        "rowIssueCount": 1,
        "rowIssueTooltip": (
            "待处理：人员身份证号尚未匹配\n"
            "需要通过人员库按姓名匹配身份证号"
        ),
        "unit": "-",
        "department": "-",
        "name": "张三",
        "identity": "待匹配",
        "insurance": "养老",
        "startMonth": "2025-08",
        "endMonth": "2026-07",
        "taskNumber": "RLSQ20260820-0001",
        "printGroup": "-",
        "printGroupId": "",
    }


def test_erp_preview_exposes_missing_dates_as_a_chinese_issue(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    view_model._on_erp_task_extraction_completed(
        {
            "summary": {
                "tasks_total": 1,
                "tasks_processed": 1,
                "tasks_failed": 0,
                "people_extracted": 1,
                "stopped": False,
            },
            "tasks": [
                {
                    "task_number": "RLSQ20260820-0002",
                    "department": "项目管理公司",
                    "parse_status": {"code": "SUCCESS", "message": "处理成功"},
                    "extraction": {"people": [{"name": "施瀛博"}]},
                }
            ],
            "rights_statement_requests": [
                {
                    "task_number": "RLSQ20260820-0002",
                    "name": "施瀛博",
                    "social_security_number": None,
                    "start_month": None,
                    "end_month": None,
                    "needs_review": True,
                    "review_reasons": ["原文月份表达不明确"],
                    "warnings": [],
                }
            ],
        }
    )

    assert application is not None
    assert view_model.records[0]["startMonth"] == "待确认"
    assert view_model.records[0]["endMonth"] == "待确认"
    codes = [item["code"] for item in view_model.recordIssues]
    assert codes == [
        "AI_DATE_MISSING",
        "AI_REVIEW_REQUIRED",
        "IDENTITY_MATCH_PENDING",
    ]
    assert view_model.recordIssues[0]["personName"] == "施瀛博"
    assert "3 项待处理" in view_model.recordStatusLabel
    assert view_model.records[0]["rowStatus"] == "error"
    assert view_model.records[0]["rowIssueCount"] == 3


def test_model_warning_is_information_not_an_actionable_issue(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    view_model._on_erp_task_extraction_completed(
        {
            "summary": {
                "tasks_total": 1,
                "tasks_processed": 1,
                "tasks_failed": 0,
                "people_extracted": 1,
                "stopped": False,
            },
            "tasks": [{"task_number": "RLSQ-003", "department": "技术中心"}],
            "rights_statement_requests": [
                {
                    "task_number": "RLSQ-003",
                    "name": "张三",
                    "social_security_number": "320101199001011234",
                    "identity_match": {
                        "code": "SUCCESS",
                        "company": "测试建设有限公司",
                        "department": "第一工程部",
                    },
                    "start_month": "2026-08",
                    "end_month": "2026-08",
                    "needs_review": False,
                    "review_reasons": [],
                    "warnings": ["非阻断性提示"],
                }
            ],
        }
    )

    assert application is not None
    assert view_model.recordIssueCount == 0
    assert view_model.recordDetailCount == 1
    assert view_model.recordIssues[0]["level"] == "info"
    assert view_model.recordIssues[0]["levelLabel"] == "信息"
    assert view_model.records[0]["unit"] == "测试建设有限公司"
    assert view_model.records[0]["department"] == "第一工程部"
    assert view_model.records[0]["rowStatus"] == "info"
    assert "非阻断性提示" in view_model.records[0]["rowIssueTooltip"]
    confirmations: list[bool] = []
    view_model.confirmationReady.connect(lambda: confirmations.append(True))

    view_model.prepareExecution()

    assert confirmations == [True]


def test_model_review_marks_only_the_related_row_yellow(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    view_model._on_erp_task_extraction_completed(
        {
            "summary": {
                "tasks_total": 1,
                "tasks_processed": 1,
                "tasks_failed": 0,
                "people_extracted": 1,
                "stopped": False,
            },
            "tasks": [{"task_number": "RLSQ-REVIEW"}],
            "rights_statement_requests": [
                {
                    "task_number": "RLSQ-REVIEW",
                    "name": "李四",
                    "social_security_number": "320101199002021235",
                    "identity_match": {
                        "code": "SUCCESS",
                        "company": "测试公司",
                        "department": "项目部",
                    },
                    "start_month": "2025-07",
                    "end_month": "2026-07",
                    "needs_review": True,
                    "review_reasons": ["原文时间表达存在两种可能解释"],
                    "warnings": [],
                }
            ],
        }
    )

    assert application is not None
    assert view_model.records[0]["rowStatus"] == "warning"
    assert view_model.records[0]["rowStatusLabel"] == "待复核"
    assert view_model.records[0]["rowIssueCount"] == 1
    assert "原文时间表达存在两种可能解释" in (
        view_model.records[0]["rowIssueTooltip"]
    )

    issue_id = str(view_model.recordIssues[0]["issueId"])
    view_model.confirmReviewIssue(issue_id)

    assert view_model.recordIssueCount == 0
    assert view_model.records[0]["rowStatus"] == "success"
    requests = view_model._erp_task_result["rights_statement_requests"]
    assert requests[0]["review_reasons"] == []
    assert requests[0]["manual_review_confirmed"] is True


def test_same_review_reason_is_shown_once_per_print_group(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    requests = []
    for name, identity in (
        ("张三", "320101199001011234"),
        ("李四", "320101199002021235"),
    ):
        requests.append(
            {
                "task_number": "RLSQ-REVIEW-GROUP",
                "group_id": "RLSQ-REVIEW-GROUP-G01",
                "group_sequence": 1,
                "group_people_count": 2,
                "source_print_mode": "combined",
                "resolved_print_mode": "combined",
                "name": name,
                "social_security_number": identity,
                "identity_match": {
                    "code": "SUCCESS",
                    "company": "测试公司",
                    "department": "项目部",
                },
                "insurance_type": "养老",
                "start_month": "2025-08",
                "end_month": "2026-07",
                "needs_review": True,
                "review_reasons": ["请确认模型拆分的打印组是否正确"],
                "warnings": [],
            }
        )

    view_model._on_erp_task_extraction_completed(
        {
            "summary": {
                "tasks_total": 1,
                "tasks_processed": 1,
                "tasks_failed": 0,
                "people_extracted": 2,
                "stopped": False,
            },
            "tasks": [{"task_number": "RLSQ-REVIEW-GROUP"}],
            "rights_statement_requests": requests,
        }
    )

    assert application is not None
    reviews = [
        issue
        for issue in view_model.recordIssues
        if issue["code"] == "AI_REVIEW_REQUIRED"
    ]
    assert len(reviews) == 1
    assert reviews[0]["personName"] == "组1"

    view_model.confirmReviewIssue(str(reviews[0]["issueId"]))

    assert view_model.recordIssueCount == 0
    assert all(not item["review_reasons"] for item in requests)


def test_erp_print_groups_are_previewed_and_unresolved_mode_can_be_selected(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    def request(
        *,
        name: str,
        identity: str,
        group_id: str,
        group_sequence: int,
        people_count: int,
        source_mode: str | None,
        resolved_mode: str | None,
    ) -> dict[str, object]:
        return {
            "task_number": "RLSQ-GROUPS",
            "group_id": group_id,
            "group_sequence": group_sequence,
            "group_people_count": people_count,
            "source_print_mode": source_mode,
            "resolved_print_mode": resolved_mode,
            "name": name,
            "social_security_number": identity,
            "identity_match": {
                "code": "SUCCESS",
                "company": "测试单位",
                "department": "项目部",
            },
            "insurance_type": "养老",
            "start_month": "2025-08",
            "end_month": "2026-07",
            "needs_review": False,
            "review_reasons": [],
            "warnings": [],
        }

    view_model._on_erp_task_extraction_completed(
        {
            "summary": {
                "tasks_total": 1,
                "tasks_processed": 1,
                "tasks_failed": 0,
                "people_extracted": 3,
                "stopped": False,
            },
            "tasks": [{"task_number": "RLSQ-GROUPS"}],
            "rights_statement_requests": [
                request(
                    name="张三",
                    identity="320101199001011234",
                    group_id="RLSQ-GROUPS-G01",
                    group_sequence=1,
                    people_count=1,
                    source_mode="combined",
                    resolved_mode="combined",
                ),
                request(
                    name="张三",
                    identity="320101199001011234",
                    group_id="RLSQ-GROUPS-G02",
                    group_sequence=2,
                    people_count=2,
                    source_mode=None,
                    resolved_mode=None,
                ),
                request(
                    name="李四",
                    identity="320101199002021235",
                    group_id="RLSQ-GROUPS-G02",
                    group_sequence=2,
                    people_count=2,
                    source_mode=None,
                    resolved_mode=None,
                ),
            ],
        }
    )

    assert application is not None
    assert view_model.erpRecordSource
    assert view_model.peopleCount == 3
    assert view_model.uniquePeopleCount == 2
    assert view_model.conditionCount == 2
    assert view_model.expectedPdfCount == 1
    assert view_model.recordIssueCount == 1
    assert view_model.recordIssues[0]["code"] == "AI_PRINT_MODE_REQUIRED"
    assert [item["label"] for item in view_model.printGroups] == ["组1", "组2"]
    assert [item["taskNumber"] for item in view_model.printGroups] == [
        "RLSQ-GROUPS",
        "RLSQ-GROUPS",
    ]
    assert view_model.printGroups[1]["modeRequired"]

    view_model.setPrintGroupMode("RLSQ-GROUPS-G02", "combined")

    assert view_model.recordIssueCount == 0
    assert view_model.expectedPdfCount == 2
    assert view_model.printGroups[1]["resolvedMode"] == "combined"
    assert not view_model.printGroups[1]["modeRequired"]


def test_preview_edit_updates_person_and_synchronizes_group_conditions(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    requests = []
    for name, identity in (
        ("张三", "320101199001011234"),
        ("李四", "320101199002021235"),
    ):
        requests.append(
            {
                "task_number": "RLSQ-EDIT",
                "group_id": "RLSQ-EDIT-G01",
                "group_sequence": 1,
                "group_people_count": 2,
                "source_print_mode": "combined",
                "resolved_print_mode": "combined",
                "name": name,
                "social_security_number": identity,
                "identity_match": {
                    "code": "SUCCESS",
                    "company": "原单位",
                    "department": "原部门",
                },
                "insurance_type": "养老",
                "start_month": "2025-08",
                "end_month": "2026-07",
                "needs_review": False,
                "review_reasons": [],
                "warnings": [],
            }
        )
    view_model._on_erp_task_extraction_completed(
        {
            "summary": {
                "tasks_total": 1,
                "tasks_processed": 1,
                "tasks_failed": 0,
                "people_extracted": 2,
                "print_groups_extracted": 1,
                "stopped": False,
            },
            "tasks": [{"task_number": "RLSQ-EDIT"}],
            "rights_statement_requests": requests,
        }
    )

    saved = view_model.updateRecord(
        2,
        "新单位",
        "新部门",
        "张三改",
        "320101199001011234",
        "工伤",
        "2026-01",
        "2026-08",
    )

    assert application is not None
    assert saved
    assert view_model.records[0]["unit"] == "新单位"
    assert view_model.records[0]["department"] == "新部门"
    assert view_model.records[0]["name"] == "张三改"
    assert view_model.records[0]["insurance"] == "工伤"
    assert view_model.records[1]["unit"] == "原单位"
    assert view_model.records[1]["insurance"] == "工伤"
    assert view_model.records[1]["startMonth"] == "2026-01"
    assert view_model.records[1]["endMonth"] == "2026-08"
    assert view_model.recordIssueCount == 0
    assert requests[0]["identity_match"]["source"] == "manual_edit"
    assert requests[1]["insurance_type"] == "工伤"


def test_preview_edit_rejects_invalid_identity_and_execution_revalidates(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["单位", "部门", "姓名", "身份证", "险种", "开始时间", "结束时间", "任务编号"])
    sheet.append(["单位", "部门", "张三", "320101199001011234", "养老", "2025-01", "2025-06", "RLSQ-001"])
    path = tmp_path / "edit.xlsx"
    workbook.save(path)
    view_model.importExcel(QUrl.fromLocalFile(str(path)))
    failures: list[tuple[str, str]] = []
    view_model.validationFailed.connect(
        lambda summary, details: failures.append((summary, details))
    )

    saved = view_model.updateRecord(
        2,
        "单位",
        "部门",
        "张三",
        "错误身份证",
        "养老",
        "2025-01",
        "2025-06",
    )

    assert application is not None
    assert not saved
    assert failures[-1][0] == "修改内容校验失败"
    assert "身份证格式错误" in failures[-1][1]

    view_model._records = [
        replace(view_model._records[0], start_month="错误月份")
    ]
    view_model.prepareExecution()

    assert failures[-1][0] == "数据校验失败"
    assert "开始时间格式错误" in failures[-1][1]


def test_template_save_dialog_has_a_default_filename(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    assert application is not None
    assert view_model.templateDefaultUrl.fileName() == "单位权益单导入模板.xlsx"


def test_manual_erp_file_selection_validates_before_opening_confirmation(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    pdf = tmp_path / "单位权益单.pdf"
    pdf.write_bytes(b"%PDF-1.7\ncontent\n%%EOF")
    ready: list[bool] = []
    view_model.erpFileReady.connect(lambda: ready.append(True))

    view_model.selectErpUploadFile(QUrl.fromLocalFile(str(pdf)))

    assert application is not None
    assert view_model.erpFileSelected
    assert view_model.erpFileName == "单位权益单.pdf"
    assert view_model.erpFileDetails.startswith("PDF · ")
    assert view_model.erpUploadStatus == "文件校验通过，请填写任务编号"
    assert ready == [True]


def test_qml_main_window_loads_with_explicit_backend(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    view_model._rights_connection_status = (
        "人工验证：验证码操作过于频繁（errorCode=12），已停止自动重试；"
        "请等待页面允许后再手动验证，这是一段用于检查布局的长状态信息"
    )
    engine = QQmlApplicationEngine()
    engine.rootContext().setContextProperty("appBackend", view_model)
    engine.setInitialProperties({"backend": view_model})
    qml_file = Path("src/ehrm/gui/qml/Main.qml").resolve()

    engine.load(QUrl.fromLocalFile(str(qml_file)))
    application.processEvents()

    assert len(engine.rootObjects()) == 1
    window = engine.rootObjects()[0]
    assert window.title() == "信息化人力工作台"
    assert window.property("backend") is view_model
    query_dialog = window.findChild(QObject, "erpTaskQueryDialog")
    assert query_dialog is not None
    query_dialog.open()
    application.processEvents()
    assert query_dialog.property("opened")
    query_dialog.close()
    record_edit_dialog = window.findChild(QObject, "recordEditDialog")
    assert record_edit_dialog is not None
    rights_settings_pane = window.findChild(QObject, "rightsSettingsPane")
    assert rights_settings_pane is not None
    assert window.findChild(QObject, "nocobaseSettingsPane") is not None
    assert window.findChild(QQuickItem, "nocobaseTestConnectionButton") is not None
    assert (
        rights_settings_pane.property("contentHeight")
        > rights_settings_pane.property("height")
    )
    status_text = window.findChild(QQuickItem, "rightsConnectionStatusText")
    test_button = window.findChild(QQuickItem, "rightsTestConnectionButton")
    assert status_text is not None
    assert test_button is not None
    status_right = status_text.mapToScene(
        QPointF(status_text.width(), 0)
    ).x()
    button_left = test_button.mapToScene(QPointF(0, 0)).x()
    assert status_right <= button_left
    assert status_text.property("lineCount") <= 2
    assert window.findChild(QQuickItem, "nocobaseApplicationsPage") is not None
    assert window.findChild(QObject, "nocobaseApplicationDetailDialog") is not None
    assert window.findChild(QObject, "nocobasePrintProgressDialog") is not None
    assert not window.property("navigationCollapsed")
    window.setProperty("navigationCollapsed", True)
    assert window.property("navigationCollapsed")
    window.close()
    application.processEvents()


def test_nocobase_application_page_is_exposed_to_qml(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    result = NocoBaseRightsApplicationPage(
        records=(
            NocoBaseRightsApplication(
                application_id=384427705696256,
                code="RLSQ20260902-0002",
                status="NEW",
                title="测试社保权益单申请2",
                problem_type="social_security_rights",
                initiator_id=25,
                initiator_name="夏国玺",
                initiation_date=datetime(
                    2026, 9, 2, 7, 16, tzinfo=timezone.utc
                ),
                estimate_time=0,
                actual_time=0,
                estimate_date=None,
                actual_date=None,
            ),
        ),
        meta=NocoBasePageMeta(
            count=1,
            page=1,
            page_size=20,
            total_page=1,
            allowed_actions={"view": (384427705696256,)},
        ),
    )

    view_model._on_nocobase_applications_succeeded(result)

    assert application is not None
    assert view_model.nocobaseApplicationsCount == 1
    assert view_model.nocobaseApplicationsPage == 1
    assert view_model.nocobaseApplicationsTotalPage == 1
    assert view_model.nocobaseApplications == [
        {
            "id": "384427705696256",
            "code": "RLSQ20260902-0002",
            "status": "NEW",
            "statusLabel": "新增",
            "title": "测试社保权益单申请2",
            "problemType": "social_security_rights",
            "problemTypeLabel": "单位社保权益单",
            "initiator": "夏国玺",
            "initiationDate": "2026-09-02",
            "estimateTime": "0.00",
            "actualTime": "0.00",
            "estimateDate": "-",
            "actualDate": "-",
        }
    ]


def test_nocobase_application_page_size_reloads_first_page(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    requested_pages: list[int] = []
    notifications: list[tuple[str, str]] = []
    monkeypatch.setattr(
        view_model,
        "loadNocobaseApplications",
        lambda page=1: requested_pages.append(page),
    )
    view_model.notification.connect(
        lambda title, message: notifications.append((title, message))
    )

    view_model.setNocobaseApplicationsPageSize(10)

    assert application is not None
    assert view_model.nocobaseApplicationsPageSize == 10
    assert requested_pages == [1]

    view_model.setNocobaseApplicationsPageSize(15)

    assert view_model.nocobaseApplicationsPageSize == 10
    assert requested_pages == [1]
    assert notifications[-1][0] == "无法切换分页"


def test_nocobase_detail_maps_people_and_starts_dedicated_print(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    view_model._output_path = tmp_path / "output"
    detail = NocoBaseRightsApplicationDetail(
        application_id=384427705696256,
        code="RLSQ20260902-0002",
        status="NEW",
        title="测试社保权益单申请2",
        problem_type="social_security_rights",
        created_at=datetime(2026, 9, 2, 7, 16, tzinfo=timezone.utc),
        initiation_date=datetime(2026, 9, 2, 7, 16, tzinfo=timezone.utc),
        estimate_time=0,
        actual_time=0,
        estimate_date=None,
        actual_date=None,
        created_by_name="夏国玺",
        initiator_name="夏国玺",
        problem_description="",
        handling_method="",
        related_persons=(
            NocoBaseRelatedPerson(
                person_id=384427705696257,
                status="NEW",
                insurance_type="elderly_care",
                start_month=datetime(2025, 5, 1, tzinfo=timezone.utc),
                end_month=datetime(2026, 1, 1, tzinfo=timezone.utc),
                identity_number="410423199005124058",
                department="第十六分公司",
                name="王明明",
                company="南京南化建设有限公司",
                print_group="组1",
            ),
            NocoBaseRelatedPerson(
                person_id=384427705696258,
                status="NEW",
                insurance_type="elderly_care",
                start_month=datetime(2025, 5, 1, tzinfo=timezone.utc),
                end_month=datetime(2026, 1, 1, tzinfo=timezone.utc),
                identity_number="320124199410222629",
                department="第十五分公司",
                name="夏素平",
                company="南京南化建设有限公司",
                print_group="组1",
            ),
            NocoBaseRelatedPerson(
                person_id=384427705696259,
                status="NEW",
                insurance_type="elderly_care",
                start_month=datetime(2025, 5, 1, tzinfo=timezone.utc),
                end_month=datetime(2026, 1, 1, tzinfo=timezone.utc),
                identity_number="320101199001011234",
                department="第十五分公司",
                name="张三",
                company="南京南化建设有限公司",
            ),
        ),
        attachment_names=(),
        allowed_actions={"view": (384427705696256,)},
    )

    class FakeWorker:
        submitted: ExcelTaskRequest | None = None

        def submit(self, request: ExcelTaskRequest) -> bool:
            self.submitted = request
            return True

    worker = FakeWorker()
    view_model._worker = worker  # type: ignore[assignment]
    started: list[bool] = []
    finished: list[bool] = []
    generic_results: list[tuple[str, str, str]] = []
    view_model.nocobasePrintStarted.connect(lambda: started.append(True))
    view_model.nocobasePrintFinished.connect(lambda: finished.append(True))
    view_model.executionFinished.connect(
        lambda title, message, details: generic_results.append(
            (title, message, details)
        )
    )
    view_model._on_nocobase_application_detail_succeeded(detail)

    assert view_model.nocobaseApplicationDetail["createdBy"] == "夏国玺"
    assert view_model.nocobaseApplicationPeople[0]["insuranceLabel"] == "养老"
    assert view_model.nocobaseApplicationPeople[0]["startMonth"] == "2025-05"
    assert view_model.nocobaseApplicationPeople[0]["printGroup"] == "组1"
    assert view_model.nocobaseApplicationPeople[2]["printGroup"] == "单独打印"

    view_model.startNocobaseApplicationPrint()

    assert application is not None
    assert started == [True]
    assert view_model.nocobasePrintState == "running"
    assert view_model.nocobasePrintRunning
    assert worker.submitted is not None
    assert worker.submitted.mode is ExportMode.BATCH
    assert len(worker.submitted.groups) == 2
    assert [len(group.records) for group in worker.submitted.groups] == [2, 1]
    assert [group.mode for group in worker.submitted.groups] == [
        ExportMode.BATCH,
        ExportMode.INDIVIDUAL,
    ]
    assert {
        item.print_group_id for item in worker.submitted.groups[0].records
    } == {"384427705696256:group:组1"}
    assert worker.submitted.groups[1].first.print_group_id == (
        "384427705696256:person:384427705696259"
    )
    assert worker.submitted.groups[0].records[0].start_month == "2025-05"
    assert worker.submitted.groups[0].records[0].end_month == "2026-01"
    assert worker.submitted.groups[0].records[0].insurance_type == "养老"
    assert worker.submitted.source_excel.is_file()
    worker.submitted.output_dir.mkdir(parents=True)
    pdf = worker.submitted.output_dir / "权益单.pdf"
    pdf.write_bytes(b"%PDF-1.7\n%%EOF")
    view_model._on_completed(
        ExcelRunResult(
            mode=ExportMode.BATCH,
            total=1,
            succeeded=1,
            failed=0,
            manifest_path=worker.submitted.output_dir / "result.json",
            result_workbook_path=worker.submitted.output_dir / "result.xlsx",
            items=(ItemResult(2, True, "SUCCESS", "处理成功", pdf),),
        )
    )

    assert finished == [True]
    assert generic_results == []
    assert view_model.nocobasePrintState == "completed"
    assert view_model.lastPdfCount == 1
    assert not worker.submitted.source_excel.exists()


def test_nocobase_application_table_resizes_without_phantom_space(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    engine = QQmlApplicationEngine()
    engine.rootContext().setContextProperty("appBackend", view_model)
    engine.setInitialProperties({"backend": view_model})
    engine.load(
        QUrl.fromLocalFile(
            str(Path("src/ehrm/gui/qml/Main.qml").resolve())
        )
    )
    window = engine.rootObjects()[0]
    window.setProperty("activeModule", 2)
    window.resize(1900, 1000)
    window.show()
    QTest.qWait(60)
    application.processEvents()

    table = window.findChild(QQuickItem, "nocobaseApplicationsTableFlick")
    assert table is not None
    assert table.property("contentWidth") == pytest.approx(table.width(), abs=1)

    window.resize(1100, 700)
    QTest.qWait(60)
    application.processEvents()
    assert table.property("contentWidth") == pytest.approx(1330, abs=1)
    maximum = max(0, table.property("contentWidth") - table.width())
    assert 0 <= table.property("contentX") <= maximum

    window.close()
    application.processEvents()


def test_double_clicking_preview_row_opens_record_editor(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["单位", "部门", "姓名", "身份证", "险种", "开始时间", "结束时间", "任务编号"])
    sheet.append(["测试单位", "项目部", "张三", "320101199001011234", "养老", "2025-01", "2025-06", "RLSQ-001"])
    source = tmp_path / "double-click.xlsx"
    workbook.save(source)
    view_model = _view_model(tmp_path)
    view_model.importExcel(QUrl.fromLocalFile(str(source)))
    engine = QQmlApplicationEngine()
    engine.rootContext().setContextProperty("appBackend", view_model)
    engine.setInitialProperties({"backend": view_model})
    engine.load(QUrl.fromLocalFile(str(Path("src/ehrm/gui/qml/Main.qml").resolve())))
    window = engine.rootObjects()[0]
    window.show()
    QTest.qWait(120)
    application.processEvents()

    def find_visual_item(item: QQuickItem, name: str) -> QQuickItem | None:
        if item.objectName() == name:
            return item
        for child in item.childItems():
            found = find_visual_item(child, name)
            if found is not None:
                return found
        return None

    preview_cell = find_visual_item(window.contentItem(), "previewCell_0_0")
    assert preview_cell is not None
    scene_point = preview_cell.mapToScene(
        QPointF(preview_cell.width() / 2, preview_cell.height() / 2)
    ).toPoint()

    QTest.mouseDClick(window, Qt.LeftButton, Qt.NoModifier, scene_point, 20)
    application.processEvents()

    editor = window.findChild(QObject, "recordEditDialog")
    assert editor is not None
    assert editor.property("opened")
    editor.close()
    window.close()
    application.processEvents()


def test_pdf_preview_dialog_loads_and_changes_pages(tmp_path: Path) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    pdf_path = tmp_path / "two-pages.pdf"
    writer = QPdfWriter(str(pdf_path))
    painter = QPainter(writer)
    painter.drawText(120, 160, "Page 1")
    writer.newPage()
    painter.drawText(120, 160, "Page 2")
    painter.end()

    view_model = _view_model(tmp_path)
    view_model._on_completed(
        ExcelRunResult(
            mode=ExportMode.BATCH,
            total=1,
            succeeded=1,
            failed=0,
            manifest_path=tmp_path / "result.json",
            result_workbook_path=tmp_path / "result.xlsx",
            items=(
                ItemResult(2, True, "SUCCESS", "处理成功", pdf_path),
            ),
        )
    )
    engine = QQmlApplicationEngine()
    engine.rootContext().setContextProperty("appBackend", view_model)
    engine.setInitialProperties({"backend": view_model})
    engine.load(QUrl.fromLocalFile(str(Path("src/ehrm/gui/qml/Main.qml").resolve())))
    window = engine.rootObjects()[0]
    window.show()
    dialog = window.findChild(QObject, "pdfPreviewDialog")
    assert dialog is not None

    assert QMetaObject.invokeMethod(dialog, "openPreview")
    for _ in range(60):
        QTest.qWait(20)
        application.processEvents()
        if dialog.property("totalPages") == 2:
            break

    assert dialog.property("opened")
    assert dialog.property("totalPages") == 2
    assert dialog.property("currentPageNumber") == 1
    assert QMetaObject.invokeMethod(dialog, "fitPage")
    QTest.qWait(100)
    application.processEvents()

    def find_visual_item(item: QQuickItem, name: str) -> QQuickItem | None:
        if item.objectName() == name:
            return item
        for child in item.childItems():
            found = find_visual_item(child, name)
            if found is not None:
                return found
        return None

    preview_viewport = find_visual_item(window.contentItem(), "pdfPreviewViewport")
    pdf_view = find_visual_item(window.contentItem(), "previewPdfView")
    assert preview_viewport is not None
    assert pdf_view is not None
    assert preview_viewport.property("clip")
    assert pdf_view.property("clip")
    assert pdf_view.property("leftMargin") > 0
    assert pdf_view.property("contentX") == pytest.approx(
        -pdf_view.property("leftMargin"),
        abs=2,
    )

    original_zoom = dialog.property("zoomPercent")
    scene_point = pdf_view.mapToScene(
        QPointF(pdf_view.width() / 2, pdf_view.height() / 2)
    )
    QTest.wheelEvent(
        window,
        scene_point,
        QPointF(0, 120).toPoint(),
        stateKey=Qt.ControlModifier,
    )
    QTest.qWait(100)
    application.processEvents()
    enlarged_zoom = dialog.property("zoomPercent")
    assert enlarged_zoom > original_zoom

    QTest.wheelEvent(
        window,
        scene_point,
        QPointF(0, 0).toPoint(),
        pixelDelta=QPointF(0, -60).toPoint(),
        stateKey=Qt.ControlModifier,
    )
    QTest.qWait(100)
    application.processEvents()
    assert dialog.property("zoomPercent") < enlarged_zoom

    assert QMetaObject.invokeMethod(dialog, "nextPage")
    QTest.qWait(80)
    application.processEvents()

    assert dialog.property("currentPageNumber") == 2
    dialog.close()
    window.close()
    application.processEvents()


def test_closing_pdf_preview_returns_to_nocobase_completed_dialog(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    pdf_path = tmp_path / "completed.pdf"
    writer = QPdfWriter(str(pdf_path))
    painter = QPainter(writer)
    painter.drawText(120, 160, "Completed")
    painter.end()

    view_model = _view_model(tmp_path)
    view_model._active_execution_source = "nocobase"
    view_model._progress_total = 1
    view_model._on_completed(
        ExcelRunResult(
            mode=ExportMode.BATCH,
            total=1,
            succeeded=1,
            failed=0,
            manifest_path=tmp_path / "result.json",
            result_workbook_path=tmp_path / "result.xlsx",
            items=(ItemResult(2, True, "SUCCESS", "处理成功", pdf_path),),
        )
    )
    engine = QQmlApplicationEngine()
    engine.rootContext().setContextProperty("appBackend", view_model)
    engine.setInitialProperties({"backend": view_model})
    engine.load(QUrl.fromLocalFile(str(Path("src/ehrm/gui/qml/Main.qml").resolve())))
    window = engine.rootObjects()[0]
    window.show()
    progress_dialog = window.findChild(QObject, "nocobasePrintProgressDialog")
    preview_dialog = window.findChild(QObject, "pdfPreviewDialog")
    assert progress_dialog is not None
    assert preview_dialog is not None

    assert QMetaObject.invokeMethod(progress_dialog, "open")
    application.processEvents()
    assert progress_dialog.property("opened")

    assert QMetaObject.invokeMethod(progress_dialog, "previewRequested")
    for _ in range(30):
        QTest.qWait(20)
        application.processEvents()
        if preview_dialog.property("opened"):
            break

    assert preview_dialog.property("opened")
    assert progress_dialog.property("opened")
    preview_dialog.close()
    application.processEvents()
    assert progress_dialog.property("opened")

    progress_dialog.close()
    window.close()
    application.processEvents()


def test_cancelled_result_is_presented_separately_from_failures(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    messages: list[tuple[str, str, str]] = []
    view_model.executionFinished.connect(
        lambda title, message, details: messages.append((title, message, details))
    )
    view_model._set_running(True)
    result = ExcelRunResult(
        mode=ExportMode.INDIVIDUAL,
        total=2,
        succeeded=1,
        failed=1,
        manifest_path=tmp_path / "result.json",
        result_workbook_path=tmp_path / "result.xlsx",
        items=(
            ItemResult(2, True, "SUCCESS", "处理成功"),
            ItemResult(3, False, "TASK_CANCELLED", "用户提前停止任务"),
        ),
    )

    view_model._on_completed(result)

    assert application is not None
    assert not view_model.running
    assert view_model.statusText == "任务已停止：成功 1，未处理 1，其他失败 0"
    assert messages[0][0] == "任务已停止"
    assert messages[0][1] == "已完成 1 人，未处理 1 人"


def test_completed_result_displays_specific_failure_reason(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    messages: list[tuple[str, str, str]] = []
    view_model.executionFinished.connect(
        lambda title, message, details: messages.append((title, message, details))
    )
    result = ExcelRunResult(
        mode=ExportMode.BATCH,
        total=1,
        succeeded=0,
        failed=1,
        manifest_path=tmp_path / "result.json",
        result_workbook_path=tmp_path / "result.xlsx",
        items=(
            ItemResult(
                2,
                False,
                "RIGHTS_API_REQUEST_FAILED",
                "智慧人社业务接口请求失败\n权益单打印接口响应超时，请稍后再试",
            ),
        ),
    )

    view_model._on_completed(result)

    assert application is not None
    assert "失败原因" in messages[0][2]
    assert "权益单打印接口响应超时" in messages[0][2]


def test_completed_result_exposes_unique_existing_pdf_files(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    pdf_path = tmp_path / "批量权益单.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n%%EOF")
    missing_path = tmp_path / "已删除.pdf"
    notifications: list[tuple[str, str]] = []
    view_model.notification.connect(
        lambda title, message: notifications.append((title, message))
    )
    result = ExcelRunResult(
        mode=ExportMode.BATCH,
        total=3,
        succeeded=2,
        failed=1,
        manifest_path=tmp_path / "result.json",
        result_workbook_path=tmp_path / "result.xlsx",
        items=(
            ItemResult(2, True, "SUCCESS", "处理成功", pdf_path),
            ItemResult(3, True, "SUCCESS", "处理成功", pdf_path),
            ItemResult(4, False, "FILE_VALIDATION_ERROR", "失败", missing_path),
        ),
    )

    view_model._on_completed(result)

    assert application is not None
    assert view_model.hasPreviewablePdfs
    assert view_model.lastPdfCount == 1
    assert view_model.lastPdfFiles[0]["name"] == "批量权益单.pdf"
    assert view_model.lastPdfFiles[0]["path"] == str(pdf_path.resolve())
    assert view_model.preparePdfPreview()

    pdf_path.unlink()

    assert not view_model.preparePdfPreview()
    assert not view_model.hasPreviewablePdfs
    assert notifications[-1][0] == "无法预览 PDF"


def test_settings_are_persisted_and_applied_to_automation(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    output = tmp_path / "downloads"
    output.mkdir()
    view_model = _view_model(tmp_path)

    view_model.setOutputFolder(QUrl.fromLocalFile(str(output)))
    view_model.setExportMode("batch")
    view_model.setBatchSize(300)
    view_model.setUploadToErp(True)
    view_model.setOpenOutputFolderAfterRun(True)
    view_model.setExecutionSpeed("stable")
    view_model.setNoResultConfirmSeconds(15)
    view_model.setPreviewDownloadDelayMs(2000)
    view_model.setDownloadTimeoutSeconds(60)

    assert application is not None
    assert view_model.outputPath == str(output)
    assert view_model.exportMode == "batch"
    assert view_model.batchSize == 300
    assert view_model.uploadToErp
    assert view_model.openOutputFolderAfterRun
    assert view_model.executionSpeed == "stable"
    assert view_model._settings.rights_statement.step_delay_ms == 1500
    assert view_model._settings.rights_statement.no_result_confirm_ms == 15_000
    assert view_model._settings.rights_statement.preview_download_delay_ms == 2000
    assert view_model._settings.rights_statement.download_timeout_ms == 60_000

    restored = _view_model(tmp_path)
    assert restored.outputPath == str(output)
    assert restored.exportMode == "batch"
    assert restored.batchSize == 300
    assert restored.uploadToErp


def test_ai_model_and_supported_reasoning_mode_are_persisted(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    assert view_model.aiModelProfile == "qwen3_5_9b"
    assert view_model.aiModelRuntimeLabel == (
        "Qwen3.5-9B（qwen3.5:9b）· 非思考"
    )
    assert [item["value"] for item in view_model.aiModelOptions] == [
        "qwen3_5_9b",
        "qwen3_8_27b",
    ]
    assert [item["value"] for item in view_model.aiReasoningOptions] == [
        "off",
        "on",
    ]

    view_model.setAiModelProfile("qwen3_8_27b")
    view_model.setAiReasoningMode("medium")

    assert application is not None
    assert view_model.aiModelProfile == "qwen3_8_27b"
    assert view_model.aiReasoningMode == "medium"
    assert view_model.aiModelRuntimeLabel == (
        "Qwen3.8-27B（qwen3.8:27b）· 中等强度"
    )
    assert [item["value"] for item in view_model.aiReasoningOptions] == [
        "off",
        "low",
        "medium",
        "max",
    ]

    restored = _view_model(tmp_path)
    assert restored.aiModelProfile == "qwen3_8_27b"
    assert restored.aiReasoningMode == "medium"
    assert restored._settings.ai.model == "qwen3.8:27b"


def test_erp_password_is_saved_in_sqlite_account_database(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    class FakeCredentialStore:
        saved: tuple[str, str] | None = None

        def load_password(self, username: str) -> str | None:
            if self.saved is not None and self.saved[0] == username:
                return self.saved[1]
            return None

        def save_password(self, username: str, password: str) -> None:
            self.saved = (username, password)

        def delete_password(self, username: str) -> None:
            pass

    store = FakeCredentialStore()
    view_model._credential_store = store

    view_model.saveErpAccount("erp-user", "secret-value")

    assert application is not None
    assert store.saved == ("erp-user", "secret-value")
    assert view_model.erpUsername == "erp-user"
    assert view_model.erpPasswordStored
    assert view_model.loadSavedErpPassword("erp-user") == "secret-value"
    assert view_model.loadSavedErpPassword("another-user") == ""


def test_rights_password_is_saved_in_sqlite_account_database(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    class FakeCredentialStore:
        saved: tuple[str, str] | None = None

        def load_password(self, username: str) -> str | None:
            if self.saved is not None and self.saved[0] == username:
                return self.saved[1]
            return None

        def save_password(self, username: str, password: str) -> None:
            self.saved = (username, password)

        def delete_password(self, username: str) -> None:
            self.saved = None

    store = FakeCredentialStore()
    view_model._rights_credential_store = store

    view_model.saveRightsAccount(
        "91320000TEST000001",
        "13800000000",
        "rights-secret",
    )

    assert application is not None
    assert store.saved == (
        "91320000TEST000001|13800000000",
        "rights-secret",
    )
    assert view_model.rightsCreditCode == "91320000TEST000001"
    assert view_model.rightsMobile == "13800000000"
    assert view_model.rightsPasswordStored
    assert view_model.rightsConnectionStatus == "账号已保存，尚未测试连接"
    assert not view_model.rightsConnectionSuccess
    assert view_model.loadSavedRightsPassword(
        "91320000TEST000001", "13800000000"
    ) == "rights-secret"


def test_nocobase_password_is_saved_in_sqlite_account_database(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)

    view_model.saveNocobaseAccount("nocobase-user", "nocobase-secret")

    assert application is not None
    assert view_model.nocobaseAccount == "nocobase-user"
    assert view_model.nocobasePasswordStored
    assert view_model.loadSavedNocobasePassword("nocobase-user") == (
        "nocobase-secret"
    )
    assert view_model.loadSavedNocobasePassword("another-user") == ""
    assert view_model.nocobaseConnectionStatus == "账号已保存，尚未测试连接"
    assert not view_model.nocobaseConnectionSuccess


def test_saved_accounts_are_restored_from_sqlite_not_preferences(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    view_model.saveErpAccount("erp-sqlite-user", "erp-sqlite-password")
    view_model.saveRightsAccount(
        "91320000SQLITE0001",
        "13900000000",
        "rights-sqlite-password",
    )
    view_model.saveNocobaseAccount(
        "nocobase-sqlite-user",
        "nocobase-sqlite-password",
    )

    restored = _view_model(tmp_path)

    assert application is not None
    assert restored.erpUsername == "erp-sqlite-user"
    assert restored.loadSavedErpPassword("erp-sqlite-user") == (
        "erp-sqlite-password"
    )
    assert restored.rightsCreditCode == "91320000SQLITE0001"
    assert restored.rightsMobile == "13900000000"
    assert restored.loadSavedRightsPassword(
        "91320000SQLITE0001", "13900000000"
    ) == "rights-sqlite-password"
    assert restored.nocobaseAccount == "nocobase-sqlite-user"
    assert restored.loadSavedNocobasePassword("nocobase-sqlite-user") == (
        "nocobase-sqlite-password"
    )
    preferences_path = (
        restored._settings.browser.user_data_dir.parent / "preferences.json"
    )
    if preferences_path.exists():
        text = preferences_path.read_text(encoding="utf-8")
        assert "erp-sqlite-user" not in text
        assert "91320000SQLITE0001" not in text
        assert "nocobase-sqlite-user" not in text


def test_clear_nocobase_login_state_keeps_credentials_and_removes_session(
    tmp_path: Path,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    view_model.saveNocobaseAccount("nocobase-user", "nocobase-secret")
    account = view_model._nocobase_credential_store.default_account()
    assert account is not None
    view_model._nocobase_credential_store.repository.save_session(
        account.id,
        "test-nocobase-token",
        expires_at=2_000_000_000,
    )
    notifications: list[tuple[str, str]] = []
    view_model.notification.connect(
        lambda title, details: notifications.append((title, details))
    )

    view_model.clearNocobaseLoginState()

    assert application is not None
    assert (
        view_model._nocobase_credential_store.repository.get_session(account.id)
        is None
    )
    assert view_model.loadSavedNocobasePassword("nocobase-user") == (
        "nocobase-secret"
    )
    assert view_model.nocobaseConnectionStatus == "NocoBase 登录状态已清除"
    assert notifications[-1][0] == "清除成功"


def test_nocobase_connection_test_persists_verified_token(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = load_settings(
        Path("config/settings.toml"),
        data_root=tmp_path / "runtime",
    )

    def encode(payload: dict[str, object]) -> str:
        raw = json.dumps(payload, separators=(",", ":")).encode("utf-8")
        return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")

    token = ".".join(
        (
            encode({"alg": "HS256", "typ": "JWT"}),
            encode({"userId": 25, "temp": True, "iat": 1, "exp": 2_000_000_000}),
            "test-signature",
        )
    )
    login_result = NocoBaseLoginResult(
        user=NocoBaseUser(25, "nocobase-user", "测试用户", "erp-id"),
        token=token,
        claims=NocoBaseTokenClaims(25, True, 1, 2_000_000_000),
    )

    class FakeRequestContext:
        def dispose(self) -> None:
            pass

    class FakeRequestFactory:
        def new_context(self) -> FakeRequestContext:
            return FakeRequestContext()

    class FakePlaywright:
        request = FakeRequestFactory()

    class FakePlaywrightContext:
        def __enter__(self) -> FakePlaywright:
            return FakePlaywright()

        def __exit__(self, *_args: object) -> None:
            pass

    class FakeAuthClient:
        def __init__(self, *_args: object) -> None:
            pass

        def sign_in(self, credentials: NocoBaseCredentials) -> NocoBaseLoginResult:
            assert credentials == NocoBaseCredentials(
                "nocobase-user",
                "nocobase-secret",
            )
            return login_result

    monkeypatch.setattr(
        nocobase_connection_module,
        "sync_playwright",
        lambda: FakePlaywrightContext(),
    )
    monkeypatch.setattr(
        nocobase_connection_module,
        "NocoBaseAuthClient",
        FakeAuthClient,
    )
    worker = NocoBaseConnectionWorker(
        settings,
        logging.getLogger("test.nocobase-connection"),
        NocoBaseCredentials("nocobase-user", "nocobase-secret"),
    )
    succeeded: list[bool] = []
    worker.succeeded.connect(lambda: succeeded.append(True))

    worker.run()

    account = AuthenticationRepository(
        settings.auth_database_path
    ).get_default_account(SystemType.NOCOBASE)
    assert succeeded == [True]
    assert account is not None
    assert account.account == "nocobase-user"
    assert account.password == "nocobase-secret"
    assert account.display_name == "测试用户"
    session = AuthenticationRepository(
        settings.auth_database_path
    ).get_session(account.id)
    assert session is not None
    assert session.session_data == token
    assert session.expires_at == 2_000_000_000
    assert session.last_verified_at is not None


def test_clear_rights_login_state_keeps_credentials_and_removes_session(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    application = QGuiApplication.instance() or QGuiApplication([])
    view_model = _view_model(tmp_path)
    view_model.saveRightsAccount(
        "91320000TEST000001",
        "13800000000",
        "rights-secret",
    )
    account = view_model._rights_credential_store.default_account()
    assert account is not None
    view_model._rights_credential_store.repository.save_session(
        account.id,
        "test-access-token",
    )
    profile = view_model._settings.browser.user_data_dir
    profile.mkdir(parents=True, exist_ok=True)
    (profile / "cookie-state").write_text("old", encoding="utf-8")
    storage_state = view_model._settings.browser.storage_state_path
    storage_state.parent.mkdir(parents=True, exist_ok=True)
    storage_state.write_text("{}", encoding="utf-8")

    notifications: list[tuple[str, str]] = []
    view_model.notification.connect(
        lambda title, details: notifications.append((title, details))
    )

    view_model.clearRightsLoginState()

    assert application is not None
    assert view_model._rights_credential_store.repository.get_session(account.id) is None
    assert not storage_state.exists()
    assert profile.is_dir()
    assert list(profile.iterdir()) == []
    assert view_model.loadSavedRightsPassword(
        "91320000TEST000001", "13800000000"
    ) == "rights-secret"
    assert view_model.rightsConnectionStatus == "智慧人社登录状态已清除"
    assert notifications[-1][0] == "清除成功"
