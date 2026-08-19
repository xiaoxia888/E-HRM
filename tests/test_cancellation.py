import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ehrm.core.error_catalog import ErrorCode
from ehrm.core.settings import load_settings
from ehrm.modules.rights_statement.excel_models import (
    EmployeeRecord,
    ExportMode,
    WorkGroup,
)
from ehrm.modules.rights_statement.excel_service import ExcelRightsStatementService


def test_cancelled_task_writes_result_for_every_unprocessed_person(
    tmp_path: Path,
) -> None:
    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["单位", "部门", "姓名", "身份证", "险种", "开始时间", "结束时间"])
    sheet.append(["测试单位", "信息部", "张三", "320101199001011234", "养老", "2025-01", "2025-06"])
    sheet.append(["测试单位", "人事部", "李四", "320101199002021235", "养老", "2025-01", "2025-06"])
    workbook.save(source)
    workbook.close()

    records = (
        EmployeeRecord(2, "测试单位", "信息部", "张三", "320101199001011234", "养老", "2025-01", "2025-06"),
        EmployeeRecord(3, "测试单位", "人事部", "李四", "320101199002021235", "养老", "2025-01", "2025-06"),
    )
    groups = [WorkGroup(sequence=1, records=records)]
    settings = load_settings(
        Path("config/settings.example.toml"),
        data_root=tmp_path / "runtime",
    )
    service = ExcelRightsStatementService(
        settings,
        logging.getLogger("test.cancellation"),
        cancel_check=lambda: True,
    )

    result = service.execute_with_page(
        object(),  # cancellation occurs before the page is accessed
        groups,
        ExportMode.BATCH,
        tmp_path / "result",
        source,
    )

    assert result.succeeded == 0
    assert result.failed == 2
    assert {item.code for item in result.items} == {str(ErrorCode.TASK_CANCELLED)}
    assert result.result_workbook_path is not None
    exported = load_workbook(result.result_workbook_path, read_only=True)
    try:
        assert exported.active.cell(2, 8).value == "用户提前停止任务"
        assert exported.active.cell(3, 8).value == "用户提前停止任务"
    finally:
        exported.close()
