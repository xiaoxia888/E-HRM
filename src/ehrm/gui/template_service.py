from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ehrm.modules.rights_statement.excel_models import EmployeeRecord


class RightsStatementTemplateService:
    HEADERS = (
        "任务编号",
        "单位",
        "部门",
        "姓名",
        "身份证",
        "险种",
        "开始时间",
        "结束时间",
    )
    INSURANCE_OPTIONS = ("养老", "工伤", "失业")

    def write(self, destination: Path) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "人员导入"

        header_fill = PatternFill("solid", fgColor="1677FF")
        for column, header in enumerate(self.HEADERS, start=1):
            cell = sheet.cell(row=1, column=column, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:H1"
        sheet.row_dimensions[1].height = 26
        widths = {
            "A": 24,
            "B": 24,
            "C": 18,
            "D": 14,
            "E": 24,
            "F": 13,
            "G": 16,
            "H": 16,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        for row in range(2, 1002):
            sheet.cell(row=row, column=1).number_format = "@"
            sheet.cell(row=row, column=5).number_format = "@"
            sheet.cell(row=row, column=7).number_format = "yyyy-mm"
            sheet.cell(row=row, column=8).number_format = "yyyy-mm"

        insurance_validation = DataValidation(
            type="list",
            formula1='"养老,工伤,失业"',
            allow_blank=False,
            error="险种只能选择养老、工伤或失业",
            errorTitle="险种填写错误",
        )
        insurance_validation.promptTitle = "请选择险种"
        insurance_validation.prompt = "养老、工伤、失业"
        insurance_validation.showInputMessage = True
        insurance_validation.showErrorMessage = True
        sheet.add_data_validation(insurance_validation)
        insurance_validation.add("F2:F1001")

        invalid_period_fill = PatternFill("solid", fgColor="FDE9E7")
        sheet.conditional_formatting.add(
            "H2:H1001",
            FormulaRule(formula=["AND(G2<>\"\",H2<>\"\",G2>H2)"], fill=invalid_period_fill),
        )

        guide = workbook.create_sheet("填写说明")
        guide_rows = (
            ("字段", "填写要求"),
            ("任务编号", "必填，ERP 人力资源事务申请编号，例如 RLSQ20260819-0001"),
            ("单位", "必填，填写参保单位名称"),
            ("部门", "必填，用于结果文件归类"),
            ("姓名", "必填"),
            ("身份证", "必填，15位或18位；本列已设置为文本格式"),
            ("险种", "必填，只能从养老、工伤、失业中选择"),
            ("开始时间", "必填，建议格式 YYYY-MM，例如 2025-05"),
            ("结束时间", "必填，不能早于开始时间"),
        )
        for row in guide_rows:
            guide.append(row)
        for cell in guide[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        guide.column_dimensions["A"].width = 18
        guide.column_dimensions["B"].width = 58
        guide.freeze_panes = "A2"

        workbook.save(destination)
        workbook.close()
        return destination

    def write_records(
        self,
        destination: Path,
        records: list[EmployeeRecord],
    ) -> Path:
        """Writes ERP-derived records as a compact execution source workbook."""

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ERP申请解析数据"
        header_fill = PatternFill("solid", fgColor="1677FF")
        for column, header in enumerate(self.HEADERS, start=1):
            cell = sheet.cell(row=1, column=column, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for record in records:
            sheet.append(
                [
                    record.task_number,
                    record.unit,
                    record.department,
                    record.name,
                    record.identity_number,
                    record.insurance_type,
                    record.start_month,
                    record.end_month,
                ]
            )
            row = sheet.max_row
            sheet.cell(row=row, column=1).number_format = "@"
            sheet.cell(row=row, column=5).number_format = "@"
            sheet.cell(row=row, column=7).number_format = "@"
            sheet.cell(row=row, column=8).number_format = "@"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:H{max(1, sheet.max_row)}"
        sheet.row_dimensions[1].height = 26
        for column, width in {
            "A": 24,
            "B": 24,
            "C": 18,
            "D": 14,
            "E": 24,
            "F": 13,
            "G": 16,
            "H": 16,
        }.items():
            sheet.column_dimensions[column].width = width
        workbook.save(destination)
        workbook.close()
        return destination
