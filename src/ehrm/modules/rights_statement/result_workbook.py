from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

from ehrm.core.error_catalog import display_message
from ehrm.modules.rights_statement.excel_models import ItemResult


_FAILURE_HEADER = "失败原因"
_ERP_STATUS_HEADER = "ERP上传结果"
_ERP_FAILURE_HEADER = "ERP失败原因"
_FAILURE_FILL = PatternFill(fill_type="solid", fgColor="FDE9E7")
_SUCCESS_FILL = PatternFill(fill_type="solid", fgColor="E8F5E9")


class ResultWorkbookWriter:
    """Copies the source workbook and appends row-level execution failures."""

    def write(
        self,
        source: Path,
        output_dir: Path,
        items: list[ItemResult],
        *,
        destination: Path | None = None,
    ) -> Path:
        keep_vba = source.suffix.lower() == ".xlsm"
        workbook = load_workbook(source, keep_vba=keep_vba)
        try:
            sheet = workbook.active
            old_max_column = sheet.max_column
            failure_column = self._column(sheet, _FAILURE_HEADER)
            self._style_header(sheet, failure_column, old_max_column, _FAILURE_HEADER)
            erp_status_column = self._column(sheet, _ERP_STATUS_HEADER)
            self._style_header(sheet, erp_status_column, old_max_column, _ERP_STATUS_HEADER)
            erp_failure_column = self._column(sheet, _ERP_FAILURE_HEADER)
            self._style_header(sheet, erp_failure_column, old_max_column, _ERP_FAILURE_HEADER)
            result_columns = (
                (failure_column, _FAILURE_HEADER),
                (erp_status_column, _ERP_STATUS_HEADER),
                (erp_failure_column, _ERP_FAILURE_HEADER),
            )

            results = {item.row_number: item for item in items}
            for row_number in range(2, sheet.max_row + 1):
                for column, _ in result_columns:
                    self._copy_neighbor_style(sheet, row_number, column, old_max_column)
                cell = sheet.cell(row=row_number, column=failure_column)
                erp_status = sheet.cell(row=row_number, column=erp_status_column)
                erp_failure = sheet.cell(row=row_number, column=erp_failure_column)
                item = results.get(row_number)
                if item is None:
                    cell.value = ""
                    erp_status.value = ""
                    erp_failure.value = ""
                    continue
                if item.success:
                    cell.value = ""
                else:
                    cell.value = display_message(item.code, item.message)
                    cell.fill = copy(_FAILURE_FILL)
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )
                if item.erp_success is None:
                    erp_status.value = "未执行"
                    erp_failure.value = ""
                elif item.erp_success:
                    erp_status.value = "上传成功"
                    erp_status.fill = copy(_SUCCESS_FILL)
                    erp_failure.value = ""
                else:
                    erp_status.value = "上传失败"
                    erp_status.fill = copy(_FAILURE_FILL)
                    erp_failure.value = display_message(
                        item.erp_code or "ERP_UPLOAD_FAILED",
                        item.erp_message,
                    )
                    erp_failure.fill = copy(_FAILURE_FILL)
                    erp_failure.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )

            sheet.column_dimensions[get_column_letter(failure_column)].width = 42
            sheet.column_dimensions[get_column_letter(erp_status_column)].width = 16
            sheet.column_dimensions[get_column_letter(erp_failure_column)].width = 42
            self._extend_filters_and_tables(
                sheet,
                old_max_column,
                max(column for column, _ in result_columns),
            )
            output_dir.mkdir(parents=True, exist_ok=True)
            resolved_destination = destination or self._destination(source, output_dir)
            workbook.save(resolved_destination)
            return resolved_destination
        finally:
            workbook.close()

    @staticmethod
    def _column(sheet, header: str) -> int:
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=column).value
            if str(value).strip() == header:
                return column
        return sheet.max_column + 1

    @staticmethod
    def _style_header(
        sheet, result_column: int, old_max_column: int, header: str
    ) -> None:
        cell = sheet.cell(row=1, column=result_column)
        if old_max_column > 0 and result_column > old_max_column:
            source = sheet.cell(row=1, column=old_max_column)
            cell._style = copy(source._style)
            cell.font = copy(source.font)
            cell.fill = copy(source.fill)
            cell.border = copy(source.border)
            cell.alignment = copy(source.alignment)
            cell.protection = copy(source.protection)
        cell.value = header

    @staticmethod
    def _copy_neighbor_style(
        sheet,
        row_number: int,
        result_column: int,
        old_max_column: int,
    ) -> None:
        if result_column <= old_max_column or old_max_column < 1:
            return
        source = sheet.cell(row=row_number, column=old_max_column)
        target = sheet.cell(row=row_number, column=result_column)
        target._style = copy(source._style)
        target.number_format = "@"

    @staticmethod
    def _extend_filters_and_tables(
        sheet,
        old_max_column: int,
        last_result_column: int,
    ) -> None:
        new_letter = get_column_letter(last_result_column)
        if sheet.auto_filter.ref:
            min_col, min_row, max_col, max_row = range_boundaries(
                sheet.auto_filter.ref
            )
            if max_col == old_max_column:
                sheet.auto_filter.ref = (
                    f"{get_column_letter(min_col)}{min_row}:{new_letter}{max_row}"
                )
        for table in sheet.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if max_col == old_max_column:
                table.ref = (
                    f"{get_column_letter(min_col)}{min_row}:{new_letter}{max_row}"
                )

    @staticmethod
    def _destination(source: Path, output_dir: Path) -> Path:
        suffix = source.suffix.lower()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate = output_dir / f"{source.stem}_执行结果_{stamp}{suffix}"
        if not candidate.exists():
            return candidate
        return output_dir / (
            f"{source.stem}_执行结果_{datetime.now():%Y%m%d_%H%M%S_%f}{suffix}"
        )
