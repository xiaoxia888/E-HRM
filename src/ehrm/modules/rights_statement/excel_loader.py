from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import replace
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ehrm.core.exceptions import ExcelValidationError
from ehrm.modules.rights_statement.excel_models import (
    EmployeeRecord,
    ExportMode,
    WorkGroup,
)


REQUIRED_HEADERS = (
    "单位",
    "部门",
    "姓名",
    "身份证",
    "险种",
    "开始时间",
    "结束时间",
    "任务编号",
)
_IDENTITY_PATTERN = re.compile(r"^(?:\d{15}|\d{17}[0-9Xx])$")
_MONTH_PATTERN = re.compile(r"^(\d{4})[-/.年]?(0?[1-9]|1[0-2])(?:月)?$")
_TASK_NUMBER_PATTERN = re.compile(r"^[A-Za-z0-9_-]{1,80}$")
_INSURANCE_OPTIONS = {"养老", "工伤", "失业"}


class RightsStatementExcelLoader:
    def load(self, path: Path) -> list[EmployeeRecord]:
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ExcelValidationError("仅支持 .xlsx 或 .xlsm 文件")
        if not path.is_file():
            raise ExcelValidationError(f"Excel 文件不存在：{path}")

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelValidationError("无法读取 Excel 文件", details=str(exc)) from exc

        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            try:
                raw_headers = next(rows)
            except StopIteration as exc:
                raise ExcelValidationError("Excel 文件为空") from exc

            headers = [str(value).strip() if value is not None else "" for value in raw_headers]
            missing = [header for header in REQUIRED_HEADERS if header not in headers]
            if missing:
                raise ExcelValidationError("Excel 缺少必要列：" + "、".join(missing))
            indexes = {header: headers.index(header) for header in REQUIRED_HEADERS}

            records: list[EmployeeRecord] = []
            errors: list[str] = []
            seen: set[tuple[str, ...]] = set()
            for row_number, values in enumerate(rows, start=2):
                if all(value in (None, "") for value in values):
                    continue
                try:
                    record = self._parse_row(row_number, values, indexes)
                    dedupe_key = (
                        record.task_number,
                        record.identity_number,
                        record.insurance_type,
                        record.start_month,
                        record.end_month,
                    )
                    if dedupe_key in seen:
                        raise ValueError("与前面数据重复")
                    seen.add(dedupe_key)
                    records.append(record)
                except ValueError as exc:
                    errors.append(f"第 {row_number} 行：{exc}")

            if errors:
                displayed = errors[:20]
                if len(errors) > len(displayed):
                    displayed.append(f"另有 {len(errors) - len(displayed)} 条错误未显示")
                raise ExcelValidationError("Excel 数据校验失败", details="\n".join(displayed))
            if not records:
                raise ExcelValidationError("Excel 中没有可执行的数据")
            return records
        finally:
            workbook.close()

    def plan(
        self,
        records: list[EmployeeRecord],
        mode: ExportMode,
        batch_size: int,
    ) -> list[WorkGroup]:
        if batch_size < 1:
            raise ExcelValidationError("batch_size 必须大于 0")
        if any(record.print_group_id for record in records):
            return self._plan_print_groups(records, batch_size)
        if mode is ExportMode.INDIVIDUAL:
            return [
                WorkGroup(
                    sequence=index,
                    records=(record,),
                    mode=ExportMode.INDIVIDUAL,
                )
                for index, record in enumerate(records, start=1)
            ]

        grouped: dict[tuple[str, str, str, str], list[EmployeeRecord]] = defaultdict(list)
        for record in records:
            grouped[record.group_key].append(record)

        plans: list[WorkGroup] = []
        sequence = 1
        for group_records in grouped.values():
            for offset in range(0, len(group_records), batch_size):
                plans.append(
                    WorkGroup(
                        sequence=sequence,
                        records=tuple(group_records[offset : offset + batch_size]),
                        mode=ExportMode.BATCH,
                    )
                )
                sequence += 1
        return plans

    def validate_records(
        self,
        records: list[EmployeeRecord],
    ) -> list[EmployeeRecord]:
        """Validates and normalizes the current in-memory preview data."""

        if not records:
            raise ExcelValidationError("没有可执行的数据")
        normalized: list[EmployeeRecord] = []
        errors: list[str] = []
        seen: set[tuple[str, ...]] = set()
        for record in records:
            try:
                item = self.normalize_record(record)
                dedupe_key = (
                    item.task_number,
                    item.print_group_id,
                    item.identity_number,
                    item.insurance_type,
                    item.start_month,
                    item.end_month,
                )
                if dedupe_key in seen:
                    raise ValueError("与前面数据重复")
                seen.add(dedupe_key)
                normalized.append(item)
            except ValueError as exc:
                errors.append(f"第 {record.row_number} 行：{exc}")
        if errors:
            displayed = errors[:20]
            if len(errors) > len(displayed):
                displayed.append(f"另有 {len(errors) - len(displayed)} 条错误未显示")
            raise ExcelValidationError(
                "数据校验失败",
                details="\n".join(displayed),
            )

        grouped_conditions: dict[tuple[str, str], set[tuple[str, str, str]]] = (
            defaultdict(set)
        )
        for item in normalized:
            if item.print_group_id:
                grouped_conditions[(item.task_number, item.print_group_id)].add(
                    (item.insurance_type, item.start_month, item.end_month)
                )
        inconsistent = [
            group_id
            for (_, group_id), conditions in grouped_conditions.items()
            if len(conditions) > 1
        ]
        if inconsistent:
            raise ExcelValidationError(
                "数据校验失败",
                details="以下打印组的险种或起止月份不一致："
                + "、".join(inconsistent),
            )
        return normalized

    def normalize_record(self, record: EmployeeRecord) -> EmployeeRecord:
        unit = self._required_text(record.unit, "单位")
        department = self._required_text(record.department, "部门")
        name = self._required_text(record.name, "姓名")
        identity = self._identity(record.identity_number)
        insurance = self._required_text(record.insurance_type, "险种")
        if insurance not in _INSURANCE_OPTIONS:
            raise ValueError("险种只能选择养老、工伤或失业")
        start = self._month(record.start_month, "开始时间")
        end = self._month(record.end_month, "结束时间")
        task_number = self._required_text(record.task_number, "任务编号")
        if not _TASK_NUMBER_PATTERN.fullmatch(task_number):
            raise ValueError("任务编号只能包含字母、数字、下划线和短横线")
        if start > end:
            raise ValueError("开始时间不能晚于结束时间")
        return replace(
            record,
            unit=unit,
            department=department,
            name=name,
            identity_number=identity,
            insurance_type=insurance,
            start_month=start,
            end_month=end,
            task_number=task_number,
        )

    @staticmethod
    def _plan_print_groups(
        records: list[EmployeeRecord],
        batch_size: int,
    ) -> list[WorkGroup]:
        grouped: dict[tuple[str, str], list[EmployeeRecord]] = defaultdict(list)
        for record in records:
            if not record.print_group_id:
                raise ExcelValidationError("ERP 打印计划中存在缺少组编号的人员")
            grouped[(record.task_number, record.print_group_id)].append(record)

        plans: list[WorkGroup] = []
        sequence = 1
        for group_records in grouped.values():
            first = group_records[0]
            conditions = {
                (
                    record.insurance_type,
                    record.start_month,
                    record.end_month,
                    record.resolved_print_mode,
                )
                for record in group_records
            }
            if len(conditions) != 1:
                raise ExcelValidationError(
                    f"打印组 {first.print_group_id} 内查询条件不一致"
                )
            if first.resolved_print_mode == ExportMode.INDIVIDUAL.value:
                for record in group_records:
                    plans.append(
                        WorkGroup(
                            sequence=sequence,
                            records=(record,),
                            mode=ExportMode.INDIVIDUAL,
                        )
                    )
                    sequence += 1
                continue
            if first.resolved_print_mode == "combined":
                for offset in range(0, len(group_records), batch_size):
                    plans.append(
                        WorkGroup(
                            sequence=sequence,
                            records=tuple(
                                group_records[offset : offset + batch_size]
                            ),
                            mode=ExportMode.BATCH,
                        )
                    )
                    sequence += 1
                continue
            # A multi-person group without an explicit or user-resolved mode
            # remains visible in the preview but is intentionally not executable.
        return plans

    def _parse_row(
        self, row_number: int, values: tuple[Any, ...], indexes: dict[str, int]
    ) -> EmployeeRecord:
        def value(header: str) -> Any:
            index = indexes[header]
            return values[index] if index < len(values) else None

        unit = self._required_text(value("单位"), "单位")
        department = self._required_text(value("部门"), "部门")
        name = self._required_text(value("姓名"), "姓名")
        identity = self._identity(value("身份证"))
        insurance = self._required_text(value("险种"), "险种")
        if insurance not in _INSURANCE_OPTIONS:
            raise ValueError("险种只能选择养老、工伤或失业")
        start = self._month(value("开始时间"), "开始时间")
        end = self._month(value("结束时间"), "结束时间")
        task_number = self._required_text(value("任务编号"), "任务编号")
        if not _TASK_NUMBER_PATTERN.fullmatch(task_number):
            raise ValueError("任务编号只能包含字母、数字、下划线和短横线")
        if start > end:
            raise ValueError("开始时间不能晚于结束时间")
        return EmployeeRecord(
            row_number=row_number,
            unit=unit,
            department=department,
            name=name,
            identity_number=identity,
            insurance_type=insurance,
            start_month=start,
            end_month=end,
            task_number=task_number,
        )

    @staticmethod
    def _required_text(value: Any, field: str) -> str:
        text = str(value).strip() if value is not None else ""
        if not text:
            raise ValueError(f"{field}不能为空")
        return text

    @staticmethod
    def _identity(value: Any) -> str:
        if isinstance(value, float):
            raise ValueError("身份证必须在 Excel 中设置为文本格式")
        text = str(value).strip() if value is not None else ""
        if not _IDENTITY_PATTERN.fullmatch(text):
            raise ValueError("身份证格式错误，应为15位或18位文本")
        return text.upper()

    @staticmethod
    def _month(value: Any, field: str) -> str:
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m")
        if isinstance(value, int) and 190001 <= value <= 999912:
            text = str(value)
        else:
            text = str(value).strip() if value is not None else ""
        match = _MONTH_PATTERN.fullmatch(text)
        if not match:
            raise ValueError(f"{field}格式错误，应为 YYYY-MM")
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
